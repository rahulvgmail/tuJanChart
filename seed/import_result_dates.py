"""Import result dates from the ResultDate and board meeting sheets."""

import logging
from datetime import date, datetime
from pathlib import Path

import openpyxl

from stockpulse.extensions import get_db
from stockpulse.models.corporate_action import BoardMeeting, ResultDate
from stockpulse.models.stock import Stock

logger = logging.getLogger(__name__)

XLSX_PATH = Path(__file__).parent.parent / "refDocs" / "Copy of Q3FY21.xlsx"


def _parse_date(val) -> date | None:
    """Parse a date from various spreadsheet formats."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    try:
        s = str(val).strip()
        if not s or s.lower() in ("na", "n/a", "-", "none", ""):
            return None
        # Try common formats
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y", "%d-%b-%Y", "%d-%b-%y"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
    except Exception:
        pass
    return None


def _quarter_from_date(d: date) -> str:
    """Determine fiscal quarter string from a date."""
    month = d.month
    year = d.year
    if month <= 3:
        return f"Q4FY{year}"
    elif month <= 6:
        return f"Q1FY{year + 1}"
    elif month <= 9:
        return f"Q2FY{year + 1}"
    else:
        return f"Q3FY{year + 1}"


def import_result_dates() -> dict:
    """Import result dates from the ResultDate sheet.

    Columns (0-indexed):
    - Col 0 (A): Security code (BSE)
    - Col 1 (B): Security Name / Symbol
    - Col 2 (C): Company name
    - Col 3 (D): Last Quarter Result Date
    - Col 10 (K): Current Quarter Result Date

    Returns:
        Dict with counts.
    """
    if not XLSX_PATH.exists():
        raise FileNotFoundError(f"Spreadsheet not found at {XLSX_PATH}")

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True, read_only=True)
    ws = wb["ResultDate"]

    session = get_db()
    stats = {"result_dates": 0, "skipped": 0}

    try:
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if len(row) < 4:
                continue

            security_code = str(row[0]).strip() if row[0] else None
            if not security_code or security_code == "None":
                stats["skipped"] += 1
                continue

            # Find stock by BSE code
            stock = session.query(Stock).filter(Stock.symbol == security_code).first()
            if not stock:
                stats["skipped"] += 1
                continue

            # Last quarter result date (col D, index 3)
            last_rd = _parse_date(row[3])
            if last_rd:
                quarter = _quarter_from_date(last_rd)
                _upsert_result_date(session, stock.id, quarter, last_rd, "spreadsheet")
                stats["result_dates"] += 1

            # Current quarter result date (col K, index 10)
            if len(row) > 10:
                curr_rd = _parse_date(row[10])
                if curr_rd:
                    quarter = _quarter_from_date(curr_rd)
                    _upsert_result_date(session, stock.id, quarter, curr_rd, "spreadsheet")
                    stats["result_dates"] += 1

            if stats["result_dates"] % 500 == 0 and stats["result_dates"] > 0:
                session.commit()

        session.commit()
        logger.info(
            "Result dates import: %d dates, %d skipped",
            stats["result_dates"],
            stats["skipped"],
        )
        return stats

    except Exception:
        session.rollback()
        logger.exception("Result dates import failed")
        raise
    finally:
        session.close()
        wb.close()


def import_board_meetings() -> dict:
    """Import board meeting data from Q1BM, Q3BM, Q4BM sheets.

    Each sheet has BSE data on the left side:
    - Col 0: Security Code
    - Col 1: Company Name
    - Col 2: Industry
    - Col 3: Purpose
    - Col 4: Meeting Date
    - Col 5: Announcement DateTime
    """
    if not XLSX_PATH.exists():
        raise FileNotFoundError(f"Spreadsheet not found at {XLSX_PATH}")

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True, read_only=True)
    session = get_db()
    stats = {"meetings": 0, "skipped": 0}

    try:
        for sheet_name in ["Q1BM", "Q3BM", "Q4BM"]:
            if sheet_name not in wb.sheetnames:
                logger.warning("Sheet %s not found, skipping", sheet_name)
                continue

            ws = wb[sheet_name]
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if len(row) < 5:
                    continue

                security_code = str(row[0]).strip() if row[0] else None
                if not security_code or security_code == "None":
                    stats["skipped"] += 1
                    continue

                purpose = str(row[3]).strip() if row[3] else None
                meeting_date = _parse_date(row[4])
                announcement_date = _parse_date(row[5]) if len(row) > 5 else None

                if not meeting_date:
                    stats["skipped"] += 1
                    continue

                stock = session.query(Stock).filter(Stock.symbol == security_code).first()
                if not stock:
                    stats["skipped"] += 1
                    continue

                # Check for duplicate
                existing = (
                    session.query(BoardMeeting)
                    .filter(
                        BoardMeeting.stock_id == stock.id,
                        BoardMeeting.meeting_date == meeting_date,
                    )
                    .first()
                )
                if existing:
                    stats["skipped"] += 1
                    continue

                session.add(
                    BoardMeeting(
                        stock_id=stock.id,
                        purpose=purpose,
                        meeting_date=meeting_date,
                        announcement_date=announcement_date,
                    )
                )
                stats["meetings"] += 1

                # Also create result date if purpose mentions results
                if purpose and ("result" in purpose.lower() or "financial" in purpose.lower()):
                    quarter = _quarter_from_date(meeting_date)
                    _upsert_result_date(session, stock.id, quarter, meeting_date, "board_meeting")

                if stats["meetings"] % 500 == 0:
                    session.commit()

            session.commit()
            logger.info("Imported board meetings from %s", sheet_name)

        logger.info(
            "Board meetings import: %d meetings, %d skipped",
            stats["meetings"],
            stats["skipped"],
        )
        return stats

    except Exception:
        session.rollback()
        logger.exception("Board meetings import failed")
        raise
    finally:
        session.close()
        wb.close()


def _upsert_result_date(session, stock_id: int, quarter: str, result_date: date, source: str):
    """Insert or update a result date entry."""
    existing = (
        session.query(ResultDate)
        .filter(ResultDate.stock_id == stock_id, ResultDate.quarter == quarter)
        .first()
    )
    if existing:
        existing.result_date = result_date
        existing.source = source
    else:
        session.add(
            ResultDate(
                stock_id=stock_id,
                quarter=quarter,
                result_date=result_date,
                source=source,
            )
        )


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    from stockpulse.app import create_app

    create_app()
    print("Importing result dates...")
    rd_stats = import_result_dates()
    print(f"Result dates: {rd_stats}")
    print("Importing board meetings...")
    bm_stats = import_board_meetings()
    print(f"Board meetings: {bm_stats}")
