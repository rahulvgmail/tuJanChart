"""Import the stock universe (~1660 stocks) from the Final sheet of the spreadsheet."""

import logging
from pathlib import Path

import openpyxl

from stockpulse.extensions import get_db
from stockpulse.models.annotation import ColorClassification
from stockpulse.models.stock import Stock

logger = logging.getLogger(__name__)

XLSX_PATH = Path(__file__).parent.parent / "refDocs" / "Copy of Q3FY21.xlsx"

# Color mapping from spreadsheet codes to standard names
COLOR_MAP = {
    "G": "Green",
    "Y": "Yellow",
    "R": "Red",
    "W": "White",
    "B": "Blue",
    "Green": "Green",
    "Yellow": "Yellow",
    "Red": "Red",
    "White": "White",
    "Blue": "Blue",
}


def import_universe(force: bool = False) -> dict:
    """Import stocks from the Final sheet.

    Columns from the spreadsheet (0-indexed):
    - Col 0 (A): BSE code (used as symbol)
    - Col 1 (B): Company name / symbol
    - Col 21 (V): C Type (sector/industry)

    Color is inferred from cell background formatting if available,
    otherwise left unset.

    Returns:
        Dict with counts: {created, updated, skipped, colors_set}
    """
    if not XLSX_PATH.exists():
        raise FileNotFoundError(f"Spreadsheet not found at {XLSX_PATH}")

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True, read_only=False)
    ws = wb["Final"]

    session = get_db()
    stats = {"created": 0, "updated": 0, "skipped": 0, "colors_set": 0}

    try:
        if force:
            # Clear existing stocks (cascade will handle related records)
            session.query(Stock).delete()
            session.commit()
            logger.info("Cleared existing stocks")

        for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=False), start=3):
            # Extract values
            bse_code = row[0].value
            name_or_symbol = row[1].value

            if not bse_code or not name_or_symbol:
                stats["skipped"] += 1
                continue

            bse_code = str(bse_code).strip()
            name_or_symbol = str(name_or_symbol).strip()

            if not bse_code or not name_or_symbol:
                stats["skipped"] += 1
                continue

            # Sector from column V (index 21)
            sector = str(row[21].value).strip() if row[21].value else None

            # The BSE code is the primary symbol
            symbol = bse_code
            company_name = name_or_symbol

            # Determine NSE symbol:
            # If bse_code is alphabetic (like HGINFRA), it's likely an NSE symbol
            # If name_or_symbol is short uppercase with no spaces, use it as NSE symbol
            nse_symbol = None
            if bse_code.isalpha() and bse_code == bse_code.upper():
                nse_symbol = bse_code
            elif len(name_or_symbol) <= 20 and " " not in name_or_symbol:
                nse_symbol = name_or_symbol

            # Check if stock already exists
            existing = session.query(Stock).filter(Stock.symbol == symbol).first()

            if existing:
                if force:
                    existing.company_name = company_name
                    existing.nse_symbol = nse_symbol
                    existing.sector = sector
                    existing.is_active = True
                    stats["updated"] += 1
                else:
                    stats["skipped"] += 1
                    continue
            else:
                stock = Stock(
                    symbol=symbol,
                    nse_symbol=nse_symbol,
                    company_name=company_name,
                    sector=sector,
                    is_active=True,
                )
                session.add(stock)
                stats["created"] += 1

            # Commit in batches
            if (stats["created"] + stats["updated"]) % 200 == 0:
                session.commit()
                logger.info(
                    "Progress: %d created, %d updated",
                    stats["created"],
                    stats["updated"],
                )

        session.commit()

        # Second pass: try to extract colors from cell formatting
        _import_colors(session, ws, stats)

        logger.info(
            "Universe import complete: %d created, %d updated, %d skipped, %d colors",
            stats["created"],
            stats["updated"],
            stats["skipped"],
            stats["colors_set"],
        )
        return stats

    except Exception:
        session.rollback()
        logger.exception("Universe import failed")
        raise
    finally:
        session.close()
        wb.close()


def _import_colors(session, ws, stats: dict):
    """Try to extract color classifications from cell background colors."""
    # Theme/indexed color mapping is unreliable in openpyxl for Google Sheets exports
    # Instead, check if there's a text-based color column or pattern
    # The spreadsheet uses cell background colors, but let's check column data first

    # Look for explicit color text in columns near the end
    for row in ws.iter_rows(min_row=3, values_only=False):
        bse_code = row[0].value
        if not bse_code:
            continue

        symbol = str(bse_code).strip()
        stock = session.query(Stock).filter(Stock.symbol == symbol).first()
        if not stock:
            continue

        # Check cell fill color of the name cell (column B, index 1)
        cell = row[1]
        color_name = _extract_color_from_fill(cell)

        if color_name:
            # Check if color already set
            existing = (
                session.query(ColorClassification)
                .filter(
                    ColorClassification.stock_id == stock.id,
                    ColorClassification.is_current == True,
                )
                .first()
            )
            if not existing:
                cc = ColorClassification(
                    stock_id=stock.id,
                    color=color_name,
                    comment="Imported from spreadsheet",
                    is_current=True,
                )
                session.add(cc)
                stats["colors_set"] += 1

    session.commit()


def _extract_color_from_fill(cell) -> str | None:
    """Extract a color name from a cell's fill color."""
    try:
        fill = cell.fill
        if fill and fill.fgColor and fill.fgColor.rgb:
            rgb = str(fill.fgColor.rgb)
            if rgb in ("00000000", "FFFFFFFF", "00FFFFFF"):
                return None  # Default/white background

            # Map common RGB values to color names
            rgb_upper = rgb.upper()
            color_map = {
                "FF00FF00": "Green",
                "FF008000": "Green",
                "FF93C47D": "Green",
                "FF6AA84F": "Green",
                "FFFFFF00": "Yellow",
                "FFFFD966": "Yellow",
                "FFFFE599": "Yellow",
                "FFFF0000": "Red",
                "FFE06666": "Red",
                "FFEA9999": "Red",
                "FFCC0000": "Red",
                "FF0000FF": "Blue",
                "FF6D9EEB": "Blue",
                "FF3D85C6": "Blue",
                "FFA4C2F4": "Blue",
                "FF9FC5E8": "Blue",
            }
            return color_map.get(rgb_upper)
    except Exception:
        pass
    return None


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    from stockpulse.app import create_app

    create_app()
    result = import_universe(force=True)
    print(f"Import result: {result}")
