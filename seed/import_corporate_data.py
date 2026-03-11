"""Import ASM entries and circuit bands from the spreadsheet."""

import logging
from pathlib import Path

import openpyxl

from stockpulse.extensions import get_db
from stockpulse.models.corporate_action import ASMEntry, CircuitBand
from stockpulse.models.stock import Stock

logger = logging.getLogger(__name__)

XLSX_PATH = Path(__file__).parent.parent / "refDocs" / "Copy of Q3FY21.xlsx"


def import_asm_entries() -> dict:
    """Import ASM (Additional Surveillance Measure) data from the ASM sheet.

    The ASM sheet has two side-by-side tables:
    - Left (Long Term ASM): Cols 0-4: SR NO, Symbol, Company Name, ISIN, Stage
    - Right (Short Term ASM): Cols 6-10: SR NO, Symbol, Company Name, ISIN, Stage

    Returns:
        Dict with counts.
    """
    if not XLSX_PATH.exists():
        raise FileNotFoundError(f"Spreadsheet not found at {XLSX_PATH}")

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True, read_only=True)
    ws = wb["ASM"]

    session = get_db()
    stats = {"created": 0, "skipped": 0}

    try:
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Process left table (Long Term ASM)
            if len(row) >= 5 and row[1]:
                _process_asm_row(session, row[1], row[4], "long_term", stats)

            # Process right table (Short Term ASM)
            if len(row) >= 11 and row[7]:
                _process_asm_row(session, row[7], row[10], "short_term", stats)

        session.commit()
        logger.info(
            "ASM import: %d created, %d skipped",
            stats["created"],
            stats["skipped"],
        )
        return stats

    except Exception:
        session.rollback()
        logger.exception("ASM import failed")
        raise
    finally:
        session.close()
        wb.close()


def _process_asm_row(session, symbol_val, stage_val, asm_type: str, stats: dict):
    """Process a single ASM entry row."""
    symbol = str(symbol_val).strip() if symbol_val else None
    if not symbol:
        stats["skipped"] += 1
        return

    # Parse stage (I, II, III, IV or 1, 2, 3, 4)
    roman_map = {"I": 1, "II": 2, "III": 3, "IV": 4}
    stage_raw = str(stage_val).strip() if stage_val else None
    stage = None
    if stage_raw:
        # Handle "Stage I", "Stage II", etc.
        clean = stage_raw.replace("Stage ", "").strip()
        if clean in roman_map:
            stage = roman_map[clean]
        else:
            try:
                stage = int(clean)
            except ValueError:
                stage = None

    # Find stock by NSE symbol or BSE code
    stock = (
        session.query(Stock)
        .filter((Stock.nse_symbol == symbol) | (Stock.symbol == symbol))
        .first()
    )
    if not stock:
        stats["skipped"] += 1
        return

    # Check for existing current ASM entry
    existing = (
        session.query(ASMEntry)
        .filter(ASMEntry.stock_id == stock.id, ASMEntry.is_current == True)
        .first()
    )
    if existing:
        # Update if stage changed
        if existing.stage != stage:
            existing.is_current = False
            session.add(
                ASMEntry(
                    stock_id=stock.id,
                    stage=stage,
                    is_current=True,
                )
            )
            stats["created"] += 1
        else:
            stats["skipped"] += 1
    else:
        session.add(
            ASMEntry(
                stock_id=stock.id,
                stage=stage,
                is_current=True,
            )
        )
        stats["created"] += 1


def import_circuit_bands() -> dict:
    """Import circuit band data from the Circuit bands sheet.

    Columns (0-indexed):
    - Col 0: Symbol
    - Col 1: Series (EQ, BE, IT)
    - Col 2: Security Name
    - Col 3: Band (2, 5, 10, 20)
    - Col 4: Remarks

    Returns:
        Dict with counts.
    """
    if not XLSX_PATH.exists():
        raise FileNotFoundError(f"Spreadsheet not found at {XLSX_PATH}")

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True, read_only=True)
    ws = wb["Circuit bands"]

    session = get_db()
    stats = {"created": 0, "updated": 0, "skipped": 0}

    try:
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if len(row) < 4:
                continue

            symbol = str(row[0]).strip() if row[0] else None
            if not symbol:
                stats["skipped"] += 1
                continue

            band_val = row[3]
            if band_val is None:
                stats["skipped"] += 1
                continue

            try:
                band_pct = int(float(str(band_val)))
            except (ValueError, TypeError):
                stats["skipped"] += 1
                continue

            # Find stock by NSE symbol or BSE code
            stock = (
                session.query(Stock)
                .filter((Stock.nse_symbol == symbol) | (Stock.symbol == symbol))
                .first()
            )
            if not stock:
                stats["skipped"] += 1
                continue

            # Upsert circuit band
            existing = (
                session.query(CircuitBand)
                .filter(CircuitBand.stock_id == stock.id)
                .first()
            )
            if existing:
                existing.band_pct = band_pct
                stats["updated"] += 1
            else:
                session.add(
                    CircuitBand(
                        stock_id=stock.id,
                        band_pct=band_pct,
                    )
                )
                stats["created"] += 1

            if (stats["created"] + stats["updated"]) % 500 == 0:
                session.commit()

        session.commit()
        logger.info(
            "Circuit bands import: %d created, %d updated, %d skipped",
            stats["created"],
            stats["updated"],
            stats["skipped"],
        )
        return stats

    except Exception:
        session.rollback()
        logger.exception("Circuit bands import failed")
        raise
    finally:
        session.close()
        wb.close()


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    from stockpulse.app import create_app

    create_app()
    print("Importing ASM entries...")
    asm_stats = import_asm_entries()
    print(f"ASM: {asm_stats}")
    print("Importing circuit bands...")
    cb_stats = import_circuit_bands()
    print(f"Circuit bands: {cb_stats}")
